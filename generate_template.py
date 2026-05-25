"""Generate template.xlsx with data-validation dropdowns on every enum column.

Open the resulting file in Excel, Google Sheets, or Numbers — clicking an
enum cell shows a dropdown of valid Meta values.

Implementation: stores all dropdown option lists on a hidden "_options"
sheet and references them by range. This works across Excel, Google
Sheets, and Numbers — unlike the inline "a,b,c" formula which Sheets and
Numbers sometimes silently drop, especially when the list starts with an
empty option.

Usage:
  python generate_template.py            # writes template.xlsx
  python generate_template.py out.xlsx
"""

import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter, quote_sheetname, absolute_coordinate
from openpyxl.worksheet.datavalidation import DataValidation

from template_options import COLUMNS

SAMPLE_ROW = {
    "existing_campaign_id": "",
    "campaign_name": "Spring Launch",
    "campaign_objective": "OUTCOME_TRAFFIC",
    "buying_type": "AUCTION",
    "special_ad_categories": "NONE",
    "campaign_daily_budget_usd": "",
    "campaign_lifetime_budget_usd": "",
    "campaign_bid_strategy": "",
    "campaign_spend_cap_usd": "",
    "campaign_start_time": "",
    "campaign_stop_time": "",
    "existing_adset_id": "",
    "adset_name": "US Broad 25-54",
    "daily_budget_usd": 50,
    "lifetime_budget_usd": "",
    "bid_strategy": "LOWEST_COST_WITHOUT_CAP",
    "bid_amount_usd": "",
    "bid_roas_floor": "",
    "daily_spend_cap_usd": "",
    "lifetime_spend_cap_usd": "",
    "pacing_type": "",
    "billing_event": "IMPRESSIONS",
    "optimization_goal": "LANDING_PAGE_VIEWS",
    "destination_type": "WEBSITE",
    "start_time": "",
    "end_time": "",
    "pixel_id": "",
    "custom_event_type": "",
    "application_id": "",
    "object_store_url": "",
    "dsa_beneficiary": "",
    "dsa_payor": "",
    "saved_audience_id": "",
    "countries": "US",
    "age_min": 25,
    "age_max": 54,
    "genders": "",
    "included_custom_audience_ids": "",
    "excluded_custom_audience_ids": "",
    "page_id": "1234567890",
    "instagram_user_id": "",
    "threads_user_id": "",
    "ad_name": "Ad A - Static",
    "image_url": "https://example.com/creative-a.jpg",
    "video_id": "",
    "video_url": "",
    "primary_text": "Discover our new spring collection.",
    "headline": "Spring is here",
    "description": "Shop the drop",
    "link_url": "https://example.com/spring",
    "display_link": "example.com",
    "url_tags": "utm_source=facebook&utm_medium=cpc&utm_campaign=spring",
    "cta": "SHOP_NOW",
    "browser_addon": "",
    "phone_number": "",
    "conversion_domain": "example.com",
    "advantage_plus_creative": "",
    "adv_add_overlays": "",
    "adv_image_touchups": "",
    "adv_music": "",
    "adv_text_generation": "",
    "adv_image_animation": "",
    "adv_product_tags": "",
    "adv_relevant_comments": "",
    "adv_enhance_cta": "",
    "adv_brightness_contrast": "",
    "adv_reveal_details": "",
    "adv_spotlights": "",
    "adv_text_overlay_translation": "",
    "adv_ig_video_subtitle": "",
    "adv_profile_card": "",
}

MAX_ROWS = 500


def build(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ads"

    headers = [name for name, _ in COLUMNS]
    ws.append(headers)
    ws.append([SAMPLE_ROW.get(name, "") for name in headers])

    opts_ws = wb.create_sheet("_options")
    opts_ws.sheet_state = "hidden"

    dropdown_count = 0
    for col_idx, (name, options) in enumerate(COLUMNS, start=1):
        if not options:
            continue
        # Drop the empty-string option from the dropdown list; the cell's
        # allow_blank=True already permits leaving the cell empty.
        values = [v for v in options if v != ""]
        if not values:
            continue
        # Lay each option list down a unique column on the _options sheet.
        opt_col = dropdown_count + 1
        for row_idx, val in enumerate(values, start=1):
            opts_ws.cell(row=row_idx, column=opt_col, value=val)
        opt_letter = get_column_letter(opt_col)
        range_ref = (
            f"={quote_sheetname('_options')}!"
            f"{absolute_coordinate(opt_letter + '1')}:"
            f"{absolute_coordinate(opt_letter + str(len(values)))}"
        )

        dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
        dv.error = f"Pick a valid {name}"
        dv.errorTitle = "Invalid value"
        dv.prompt = "Choose from the list"
        dv.promptTitle = name
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS}")
        ws.add_data_validation(dv)
        dropdown_count += 1

    for col_idx, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Wrote {path} with {dropdown_count} dropdown columns (options on hidden _options sheet).")


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "template.xlsx")
